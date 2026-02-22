import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from config import ARQUIVO_EXCEL 


# --- 1. Dicionários de Configuração ---
def formatacoes():
    return {
        "contabil": '_-R$* #,##0.00_-;-R$* #,##0.00_-;_-R$* "-"??_-;_-@_-',
        "data": "dd/mm/yyyy hh:mm:ss",
        "texto": '@'
    }

                
# --- 3. Função Principal da Aba Caixa ---
def formatacoes():
    return {
        "contabil": '_-R$* #,##0.00_-;-R$* #,##0.00_-;_-R$* "-"??_-;_-@_-',
        "data": "dd/mm/yyyy hh:mm:ss",
        "texto": '@'
    }

def caixa(dados_excel, arquivo_destino):
    salvar_aba(
        "Caixa",
        dados_excel["Caixa"].copy(),
        arquivo_destino,
        mapa_caixa(),
        numericos_caixa(),
        datas_caixa()
    )

  
def mapa_caixa():
    return {
        "Valor unitario compra": "contabil",
        "Valor unitario venda": "contabil",
        "Valor recebido": "contabil",
        "Data": "data", 
        "Operacao": "texto",
        "Sexo": "texto",
        "Participante": "texto",
        "Nome do produto": "texto",
        "Desconto": "contabil",
        "Tamanho": "texto",
        "Forma Pag": "texto",
        "Parcelas": "texto"
    }

def numericos_caixa():
    return {
        "Valor unitario compra": True,
        "Valor unitario venda": True,
        "Valor recebido": True,
        "Desconto": True,
    }

def datas_caixa():
    return ["Data"]

def compras(dados_excel, arquivo_destino):
    salvar_aba(
        "Compras",
        dados_excel["Compras"].copy(),
        arquivo_destino,
        mapa_compras(),
        numericos_compras(),
        datas_compras()
    )

def mapa_compras():
    return {
        "Valor unitario compra": "contabil",
        "Data da compra": "data",
        "Nome do produto": "texto",
        "Tamanho": "texto",
        "Sexo": "texto",
        "Quantidade": "texto",
    }

def numericos_compras():
    return {
        "Valor unitario compra": True
    }

def datas_compras():
    return ["Data da compra"]


def vendas(dados_excel, arquivo_destino):
    salvar_aba(
        "Vendas",
        dados_excel["Vendas"].copy(),
        arquivo_destino,
        mapa_vendas(),
        numericos_vendas(),
        datas_vendas()
    )

def mapa_vendas():
    return {
        "Cliente": "texto",
        "Valor unitario venda": "contabil",
        "Data": "data",
        "Nome do produto": "texto",
        "Tamanho": "texto",
        "Sexo": "texto",
        "Quantidade": "texto",
    }

def numericos_vendas():
    return {
        "Valor unitario venda": True
    }

def datas_vendas():
    return ["Data"]


def estoque(dados_excel, arquivo_destino):
    salvar_aba(
        "Estoque",
        dados_excel["Estoque"].copy(),
        arquivo_destino,
        mapa_estoque(),
        numericos_estoque()
    )

def mapa_estoque():
    return {
        "Nome do produto": "texto",
        "Tamanho": "texto",
        "Quantidade": "texto",
        "Valor unitario compra": "contabil",
    }

def numericos_estoque():
    return {
        "Valor unitario compra": True
    }

def Areceber(dados_excel, arquivo_destino):
    salvar_aba(
        "A Receber",
        dados_excel["A Receber"].copy(),
        arquivo_destino,
        mapa_areceber(),
        numericos_areceber(),
        datas_areceber()
    )

def mapa_areceber():
    return {
        "Cliente": "texto",
        "Valor total": "contabil",
        "Data vencimento": "data",
        "Nome do produto": "texto",
        "Valor parcela": "contabil",
        "Parcela": "texto",
        "Valor total pendente": "contabil"
    }

def numericos_areceber():
    return {
        "Valor total": True,
        "Valor parcela": True,
        "Valor total pendente": True
    }

def datas_areceber():
    return ["Data vencimento"]


def Apagar(dados_excel, arquivo_destino):
    salvar_aba(
        "A Pagar",
        dados_excel["A Pagar"].copy(),
        arquivo_destino,
        mapa_apagar(),
        numericos_apagar(),
        datas_apagar()
    )

def mapa_apagar():
    return {
        "Cliente": "texto",
        "Valor total": "contabil",
        "Data vencimento": "data",
        "Nome do produto": "texto",
        "Valor parcela": "contabil",
        "Parcela": "texto",
        "Valor total pendente": "contabil"
    }

def numericos_apagar(): 
    return {
        "Valor total": True,
        "Valor parcela": True,
        "Valor total pendente": True
    }

def datas_apagar(): 
    return ["Data vencimento"]


def clientes(dados_excel, arquivo_destino):
    salvar_aba(
        "Clientes",
        dados_excel["Clientes"].copy(),
        arquivo_destino,
        mapa_clientes()
    )

def mapa_clientes():
    return {
        "Nome": "texto",
        "Cpf": "texto",
        "Cidade": "texto",
        "Telefone": "texto",
    }


def salvar_aba(nome_aba, df, arquivo_destino, mapa_formatacao, mapa_numerico=None, colunas_data=None):
# Converter colunas de data para datetime e remover timezone
    if colunas_data:
        for col in colunas_data:
            if col in df.columns:
             df[col] = pd.to_datetime(df[col], unit='ns', errors='coerce')
             if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None)

# Converter colunas numéricas para float
    if mapa_numerico:
        for col in mapa_numerico.keys():
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

# Salvar a aba no Excel
    with pd.ExcelWriter(arquivo_destino, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=nome_aba, index=False)
        aplicar_estilos_excel(writer, nome_aba, df, mapa_formatacao, formatacoes())


def aplicar_estilos_excel(writer, nome_aba, df, mapeamento, formatos):
    worksheet = writer.sheets[nome_aba]
    for col_num, col_nome in enumerate(df.columns, 1):
        if col_nome in mapeamento:
            chave_formato = mapeamento[col_nome]   
            mascara = formatos.get(chave_formato) 
            
            if mascara:
                letra_col = get_column_letter(col_num)
      
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_num)
                    cell.number_format = mascara
                    cell.alignment = Alignment(horizontal="center")

    
                worksheet.column_dimensions[letra_col].width = 18


##Principal
def executar_formatacao(dados_excel):

    caminho = "C:/Users/vinicius.gomes/Desktop/py/loja-roupa-py/dados/"
    arquivo_saida = caminho + "Controle.xlsx"
    formatarcaixa = caixa(dados_excel, arquivo_saida)
    formatarcompras = compras(dados_excel, arquivo_saida)
    formatarvendas = vendas(dados_excel, arquivo_saida)
    formatarestoque = estoque(dados_excel, arquivo_saida)
    formatarAreceber = Areceber(dados_excel, arquivo_saida)
    formatarApagar = Apagar(dados_excel, arquivo_saida)
    formatarclientes = clientes(dados_excel, arquivo_saida)

    return {"formcaixa": formatarcaixa,
        "formcompras": formatarcompras,
        "formvendas": formatarvendas,
        "formestoque": formatarestoque,
        "formareceber": formatarAreceber,
        "formapagar": formatarApagar,
        "formclientes": formatarclientes
        }